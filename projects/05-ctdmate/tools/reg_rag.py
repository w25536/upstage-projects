# ctdmate/tools/reg_rag.py
from __future__ import annotations
from pathlib import Path
from typing import Dict, List, Any, Optional
import re
import openpyxl

# config
try:
    from ctdmate.app import config as CFG
except Exception:
    from ..app import config as CFG  # type: ignore

# RAG + 정규화 도구
try:
    from ctdmate.rag.mfds_rag import MFDSRAGTool
    from ctdmate.rag.glossary_rag import GlossaryRAGTool
    from ctdmate.rag.term_normalizer import TermNormalizer
except Exception:
    from ..rag.mfds_rag import MFDSRAGTool  # type: ignore
    from ..rag.glossary_rag import GlossaryRAGTool  # type: ignore
    from ..rag.term_normalizer import TermNormalizer  # type: ignore

SHEET_TO_MODULE = {
    "TM_5_M2_3_QOS": "M2.3",
    "TM_5_M2_4_Nonclinical_Ove": "M2.4",
    "TM_5_M2_5_Clinical_Overvi": "M2.5",
    "TM_5_M2_6_Nonclinical_Sum": "M2.6",
    "TM_5_M2_7_Clinical_Summar": "M2.7",
    "TM_5_Admin_Labeling_KR": "M1",
    "TM_5_Nonclinical": "M2.6",
    "TM_5_Phase1": "M2.7",
    "TM_5_Phase2": "M2.7",
    "TM_5_Phase3": "M2.7",
}

def _normalize_section(s: str) -> str:
    s = (s or "").strip().upper()
    return s if s.startswith("M") else f"M{s}"

class RegulationRAGTool:
    """
    규제 검증·정규화·근거 반환.
    임계값 근거: CFG.COVERAGE_MIN, CFG.RAG_CONF_MIN, CFG.VIO_MAX, CFG.GENERATE_GATE
    """

    def __init__(
        self,
        auto_normalize: bool = True,
        max_violations: Optional[int] = None,
        coverage_threshold: Optional[float] = None,
        enable_rag: bool = True,
        llama_client=None,
    ):
        self.auto_normalize = auto_normalize
        self.max_violations = max_violations if max_violations is not None else CFG.VIO_MAX
        self.coverage_threshold = coverage_threshold if coverage_threshold is not None else CFG.COVERAGE_MIN
        self.enable_rag = enable_rag
        self.llama_client = llama_client

        self.mfds_rag: Optional[MFDSRAGTool] = None
        self.glossary_rag: Optional[GlossaryRAGTool] = None
        self.normalizer: Optional[TermNormalizer] = None
        self.combined_retriever: Optional[Any] = None  # Retriever for combined_regulations

        if enable_rag:
            try:
                self.mfds_rag = MFDSRAGTool()
                self.glossary_rag = GlossaryRAGTool()
                # Initialize combined regulations retriever (reusable for ICH/MFDS/GLOSSARY)
                try:
                    from ctdmate.rag.retriever import Retriever
                except Exception:
                    from ..rag.retriever import Retriever  # type: ignore
                self.combined_retriever = Retriever(collection="combined_regulations", use_bm25=False)
            except Exception:
                self.enable_rag = False

        if auto_normalize:
            try:
                self.normalizer = TermNormalizer(client=self.llama_client)
            except Exception:
                self.normalizer = None

    # -------- Excel 전체 검증 --------
    def validate_excel(self, excel_path: str, auto_fix: bool = True) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        results: List[Dict[str, Any]] = []
        total_violations = 0
        total_coverage = 0.0

        for sheet_name in wb.sheetnames:
            module = _normalize_section(SHEET_TO_MODULE.get(sheet_name, ""))
            if not module:
                continue

            ws = wb[sheet_name]
            content = self._extract_sheet_content(ws)
            if len(content) < 10:
                continue

            r = self.validate_and_normalize(section=module, content=content, auto_fix=auto_fix)
            r["sheet_name"] = sheet_name
            r["module"] = module
            results.append(r)

            total_violations += len(r["violations"])
            total_coverage += r["coverage"]

        validated_count = len(results)
        pass_count = sum(1 for r in results if r["pass"])
        return {
            "total_sheets": len(wb.sheetnames),
            "validated_sheets": validated_count,
            "results": results,
            "summary": {
                "total_violations": total_violations,
                "avg_coverage": (total_coverage / validated_count) if validated_count else 0.0,
                "pass_rate": (pass_count / validated_count) if validated_count else 0.0,
            },
        }

    def _extract_sheet_content(self, ws) -> str:
        lines = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append(" ".join(cells))
        return "\n".join(lines)

    # -------- 단일 섹션 검증 --------
    def validate_and_normalize(self, section: str, content: str, auto_fix: bool = True) -> Dict[str, Any]:
        section = _normalize_section(section)

        guideline_results: List[Dict[str, Any]] = []
        if self.enable_rag and self.mfds_rag:
            try:
                guideline_results = self.mfds_rag.search_by_module(query=content[:500], module=section, k=5)
            except Exception as e:
                import sys
                print(f"[DEBUG] MFDS search failed: {e}", file=sys.stderr)
                import traceback
                traceback.print_exc()
                guideline_results = []

        violations = self._detect_violations(content, guideline_results)

        normalized_content = content
        if auto_fix and violations and self.normalizer:
            try:
                normalized_content = self._normalize_content(content, violations)
            except Exception:
                normalized_content = content

        coverage = self._calculate_coverage(normalized_content, guideline_results)
        rag_conf = self._calculate_confidence(guideline_results)
        gloss = self._glossary_hit_rate(normalized_content)

        if coverage < self.coverage_threshold:
            guideline_results.extend(self._expand_coverage(section, normalized_content))
            coverage = self._calculate_coverage(normalized_content, guideline_results)

        citations = self._generate_citations(guideline_results)

        vio_w = self._violation_weight(violations)
        score_raw = 0.55 * coverage + 0.30 * rag_conf + 0.15 * gloss
        score = max(0.0, score_raw - 0.05 * vio_w)

        passed = (
            coverage >= CFG.COVERAGE_MIN and
            rag_conf >= CFG.RAG_CONF_MIN and
            vio_w <= CFG.VIO_MAX
        )

        return {
            "validated": True,
            "pass": passed,
            "violations": violations,
            "normalized_content": normalized_content,
            "coverage": coverage,
            "citations": citations,
            "rag_conf": rag_conf,
            "metrics": {
                "score": score,
                "score_raw": score_raw,
                "vio_weight": vio_w,
                "glossary_hit": gloss,
                "thresholds": {
                    "coverage_min": CFG.COVERAGE_MIN,
                    "rag_conf_min": CFG.RAG_CONF_MIN,
                    "vio_max": CFG.VIO_MAX,
                    "generate_gate": CFG.GENERATE_GATE,
                },
            },
        }

    # -------- Helpers --------
    def _detect_violations(self, content: str, guidelines: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        4-tier 위반사항 탐지 (ICH + MFDS + Glossary + 구조)

        Args:
            content: 검증할 내용
            guidelines: MFDS 가이드라인 검색 결과

        Returns:
            위반사항 리스트
        """
        violations: List[Dict[str, Any]] = []

        # 1. ICH 통합 규제 인덱스 검색
        ich_violations = self._check_ich_guidelines(content)
        violations.extend(ich_violations)

        # 2. MFDS 가이드라인 기반 검증
        mfds_violations = self._check_mfds_guidelines(content, guidelines)
        violations.extend(mfds_violations)

        # 3. 용어 표준화 체크
        term_violations = self._check_terminology(content)
        violations.extend(term_violations)

        # 4. Placeholder 체크 (기존 로직)
        low = content.lower()
        placeholders = ["tbd", "to be defined", "to be decided", "미정", "lorem ipsum", "as appropriate", "etc."]
        if any(p in low for p in placeholders):
            violations.append({
                "type": "placeholder",
                "description": "placeholder detected",
                "suggestion": "replace placeholders with actual values",
                "severity": "major",
            })

        return violations

    def _check_ich_guidelines(self, content: str) -> List[Dict[str, Any]]:
        """
        ICH 통합 규제 인덱스로 가이드라인 위반 체크

        Args:
            content: 검증할 내용

        Returns:
            ICH 기반 위반사항 리스트
        """
        violations = []

        if not self.enable_rag or not self.combined_retriever:
            return violations

        try:
            # 통합 규제 인덱스에서 ICH 가이드라인 검색 (shared retriever 사용)
            ich_results = self.combined_retriever.vector_search(
                query=content[:500],
                k=5,
                where={"metadata.source": "ICH"}
            )

            # ICH 가이드라인 기반 위반사항 탐지
            for result in ich_results:
                score = result.get("score", 0.0)
                metadata = result.get("metadata", {})

                # 낮은 유사도는 잠재적 위반으로 간주
                if score < 0.85:
                    violations.append({
                        "type": "ICH_GUIDELINE",
                        "severity": "medium" if score < 0.75 else "low",
                        "description": f"ICH {metadata.get('module', 'N/A')} 가이드라인과 낮은 일치도",
                        "suggestion": f"다음 ICH 가이드라인을 참고하세요: {metadata.get('title', 'N/A')}",
                        "source": "ICH",
                        "module": metadata.get("module", "N/A"),
                        "section": metadata.get("section", "N/A"),
                        "score": score,
                        "reference": result.get("content", "")[:200] + "..."
                    })

        except Exception:
            # ICH 검색 실패 시 무시 (선택적 기능)
            pass

        return violations

    def _check_mfds_guidelines(self, content: str, guidelines: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        MFDS 가이드라인 기반 위반사항 체크

        Args:
            content: 검증할 내용
            guidelines: MFDS 검색 결과

        Returns:
            MFDS 기반 위반사항 리스트
        """
        violations = []

        for guideline in guidelines:
            score = guideline.get("score", 0.0)
            metadata = guideline.get("metadata", {})

            # 낮은 유사도는 잠재적 위반
            if score < 0.80 and metadata.get("source") == "MFDS":
                violations.append({
                    "type": "MFDS_GUIDELINE",
                    "severity": "high" if score < 0.70 else "medium",
                    "description": f"MFDS {metadata.get('module', 'N/A')} 가이드라인과 낮은 일치도",
                    "suggestion": f"다음 MFDS 가이드라인을 참고하세요: {metadata.get('title', 'N/A')}",
                    "source": "MFDS",
                    "module": metadata.get("module", "N/A"),
                    "score": score,
                    "reference": guideline.get("content", "")[:200] + "..."
                })

        return violations

    def _check_terminology(self, content: str) -> List[Dict[str, Any]]:
        """
        용어집 기반 용어 표준화 체크

        Args:
            content: 검증할 내용

        Returns:
            용어 위반사항 리스트
        """
        violations = []

        if not self.enable_rag or not self.combined_retriever:
            return violations

        try:
            # 통합 규제 인덱스에서 GLOSSARY 검색 (shared retriever 사용)
            term_results = self.combined_retriever.vector_search(
                query=content[:300],
                k=3,
                where={"metadata.source": "GLOSSARY"}
            )

            # 비표준 용어 탐지
            for term_result in term_results:
                score = term_result.get("score", 0.0)
                metadata = term_result.get("metadata", {})

                # 낮은 유사도는 비표준 용어 사용 가능성
                if score < 0.75:
                    violations.append({
                        "type": "TERMINOLOGY",
                        "severity": "low",
                        "description": "비표준 용어 사용 가능성",
                        "suggestion": f"표준 용어 사용 권장: {metadata.get('term', 'N/A')} ({metadata.get('term_en', 'N/A')})",
                        "source": "GLOSSARY",
                        "score": score,
                        "reference": term_result.get("content", "")[:150] + "..."
                    })

        except Exception:
            pass

        return violations

    def _normalize_content(self, content: str, violations: List[Dict[str, Any]]) -> str:
        if not self.normalizer:
            return content
        return self.normalizer.normalize(content)

    def _calculate_coverage(self, content: str, guidelines: List[Dict[str, Any]]) -> float:
        """
        Coverage는 검색된 가이드라인의 **품질과 관련성**을 측정합니다.

        측정 요소:
        1. 검색 점수 (60%): 벡터 유사도 기반 관련성
        2. 가이드라인 개수 (40%): 충분한 규제 근거 확보 여부
        """
        if not guidelines:
            return 0.0

        # 1. 평균 검색 점수 (상위 5개)
        top_scores = [g.get("score", 0.0) for g in guidelines[:5]]
        avg_score = sum(top_scores) / len(top_scores) if top_scores else 0.0

        # 2. 가이드라인 충분성 (최소 3개 권장, 5개 이상 만점)
        count_score = min(1.0, len(guidelines) / 5.0)

        # 가중합산
        coverage = 0.6 * avg_score + 0.4 * count_score

        return min(1.0, max(0.0, coverage))

    def _calculate_confidence(self, results: List[Dict[str, Any]]) -> float:
        if not results:
            return 0.0
        scores = [r.get("score", 0.0) for r in results[:3]]
        return sum(scores) / len(scores) if scores else 0.0

    def _glossary_hit_rate(self, content: str) -> float:
        if not self.glossary_rag:
            return 0.0
        try:
            hits = self.glossary_rag.search(content[:120]) or []
        except Exception:
            hits = []
        return 0.0 if not hits else min(1.0, sum(float(r.get("score", 0.0)) for r in hits) / len(hits))

    def _expand_coverage(self, section: str, content: str) -> List[Dict[str, Any]]:
        if not (self.enable_rag and self.mfds_rag):
            return []
        try:
            return self.mfds_rag.search_with_mmr(query=content[:500], k=3, fetch_k=10, lambda_mult=0.5)
        except Exception:
            return []

    def _generate_citations(self, guidelines: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        cites: List[Dict[str, Any]] = []
        for r in guidelines or []:
            md = r.get("metadata", {}) or {}
            cites.append({
                "source": md.get("source", "N/A"),
                "section": md.get("module", "N/A"),
                "page": md.get("page", "N/A"),
                "snippet": (r.get("content") or "")[:200] + "...",
                "score": r.get("score", 0.0)
            })
        return cites

    def _violation_weight(self, vlist: List[Dict[str, Any]]) -> int:
        w = {"minor": 1, "major": 2, "critical": 4}
        tot = 0
        for v in vlist:
            sev = str(v.get("severity", "major")).lower()
            tot += w.get(sev, 2)
        return tot

if __name__ == "__main__":
    tool = RegulationRAGTool(auto_normalize=False, enable_rag=False)
    sample = "임상은 다기관 무작위배정 이중맹검으로 수행되었다. TBD."
    out = tool.validate_and_normalize("M2.7", sample, auto_fix=False)
    print({"pass": out["pass"], "metrics": out["metrics"]})
